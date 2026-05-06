# NOTICE: This is copyrighted material. It is not to be reused, redistributed, or used in training datasets without explicit permission from the author.
"""
Export prewalk-bundle.json to a multi-sheet Excel workbook.
One worksheet per feature kind; all sections combined.

Output: prewalk-data-export.xlsx  (same folder as this script)
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA    = os.path.dirname(os.path.abspath(__file__))
OUT_XLS = os.path.join(DATA, 'prewalk-data-export.xlsx')

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _midpoint(coords):
    """Return (lat, lng) midpoint of a LineString coordinate list [[lng,lat],...]."""
    c = coords[len(coords) // 2]
    return round(c[1], 7), round(c[0], 7)

def _pt(coords):
    """Return (lat, lng) from a single Point coordinate [lng, lat]."""
    return round(coords[1], 7), round(coords[0], 7)

def geom_point(pin):
    """Single lat/lng for Point geometries."""
    g = pin.get('geometry', {})
    if g.get('type') == 'Point':
        lat, lng = _pt(g['coordinates'])
        return {'Latitude': lat, 'Longitude': lng}
    # Fallback: midpoint of a LineString
    if g.get('type') == 'LineString' and g.get('coordinates'):
        lat, lng = _midpoint(g['coordinates'])
        return {'Latitude': lat, 'Longitude': lng}
    return {'Latitude': '', 'Longitude': ''}

def geom_line(pin):
    """Start / mid / end lat-lng for LineString geometries."""
    g = pin.get('geometry', {})
    if g.get('type') != 'LineString' or not g.get('coordinates'):
        return {}
    coords = g['coordinates']
    slat, slng = _pt(coords[0])
    elat, elng = _pt(coords[-1])
    mlat, mlng = _midpoint(coords)
    return {
        'Start Lat': slat, 'Start Lng': slng,
        'Mid Lat':   mlat, 'Mid Lng':   mlng,
        'End Lat':   elat, 'End Lng':   elng,
    }

def base(sec, pin):
    """Fields common to every pin type."""
    reported = pin.get('reported_in', [])
    return {
        'Section':          sec['id'],
        'Section Name':     sec['name'],
        'Project Code':     sec.get('project_code', ''),
        'Pin ID':           pin['id'],
        'Asset ID':         pin.get('asset_id', ''),
        'Source':           pin.get('source', ''),
        'Status':           pin.get('status', ''),
        'Reported In':      ', '.join(reported) if reported else '',
        'Stationing':       pin.get('sta', ''),
        'Stationing (ft)':  pin.get('sta_ft', ''),
    }

def a(pin, *keys):
    """Pull attrs fields as a dict, returning '' for missing values."""
    attrs = pin.get('attrs', {})
    return {k: attrs.get(k, '') for k in keys}

# ---------------------------------------------------------------------------
# Per-sheet row builders
# ---------------------------------------------------------------------------

def rows_guardrail(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID', 'Asset ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'MP Start', 'MP End', 'Side', 'Length (ft)', 'Length Source',
        'Barrier Type', 'Inspection Date',
        'Speed Limit', 'Road Grade (%)', 'Hazard Behind',
        'Crashworthy', 'Test Level', 'Repair Action', 'Repair Cost',
        'Photo',
        'Start Lat', 'Start Lng', 'Mid Lat', 'Mid Lng', 'End Lat', 'End Lng',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'guardrail':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            r.update({
                'MP Start':        at.get('mp_start', ''),
                'MP End':          at.get('mp_end', ''),
                'Side':            at.get('side', ''),
                'Length (ft)':     at.get('length_ft', ''),
                'Length Source':   at.get('length_source', ''),
                'Barrier Type':    at.get('rpt_type', ''),
                'Inspection Date': at.get('inspection_date', ''),
                'Speed Limit':     at.get('speed_limit', ''),
                'Road Grade (%)':  at.get('road_grade_pct', ''),
                'Hazard Behind':   at.get('hazard_behind', ''),
                'Crashworthy':     at.get('crashworthy', ''),
                'Test Level':      at.get('test_level', ''),
                'Repair Action':   at.get('repair_action', ''),
                'Repair Cost':     at.get('repair_cost', ''),
                'Photo':           at.get('photo', ''),
                **geom_line(pin),
            })
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_wall(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID', 'Asset ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'MP Start', 'MP End', 'Side', 'Length (ft)', 'Length Source',
        'Wall Type', 'Wall Function', 'Wall Material',
        'Rating (0–100)', 'Repair Cost', 'Inspection Date',
        'Photo',
        'Start Lat', 'Start Lng', 'Mid Lat', 'Mid Lng', 'End Lat', 'End Lng',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'wall':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            r.update({
                'MP Start':        at.get('mp_start', ''),
                'MP End':          at.get('mp_end', ''),
                'Side':            at.get('side', ''),
                'Length (ft)':     at.get('length_ft', ''),
                'Length Source':   at.get('length_source', ''),
                'Wall Type':       at.get('rpt_type', ''),
                'Wall Function':   at.get('wall_function', ''),
                'Wall Material':   at.get('wall_material', ''),
                'Rating (0–100)':  at.get('rating', ''),
                'Repair Cost':     at.get('repair_cost', ''),
                'Inspection Date': at.get('inspection_date', ''),
                'Photo':           at.get('photo', ''),
                **geom_line(pin),
            })
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_culvert(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'Culvert Type', 'Material', 'FMSS Culvert Type',
        'FMSS Loc', 'FMSS Asset',
        'Road Bearing (deg)',
        'EP1 Role', 'EP1 Lat', 'EP1 Lng', 'EP1 Headwall', 'EP1 Flange/End Sect.',
        'EP2 Role', 'EP2 Lat', 'EP2 Lng', 'EP2 Headwall', 'EP2 Flange/End Sect.',
        'Mid Lat', 'Mid Lng',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'culvert':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            ep1 = at.get('endpoint_1') or {}
            ep2 = at.get('endpoint_2') or {}
            r.update({
                'Culvert Type':         at.get('rpt_type', ''),
                'Material':             at.get('material', ''),
                'FMSS Culvert Type':    at.get('fmss_type', ''),
                'FMSS Loc':             at.get('fmss_loc', ''),
                'FMSS Asset':           at.get('fmss_asset', ''),
                'Road Bearing (deg)':   at.get('road_bearing_deg', ''),
                'EP1 Role':             ep1.get('role', ''),
                'EP1 Lat':              ep1.get('lat', ''),
                'EP1 Lng':              ep1.get('lng', ''),
                'EP1 Headwall':         ep1.get('headwall', ''),
                'EP1 Flange/End Sect.': ep1.get('flange_end_section', ''),
                'EP2 Role':             ep2.get('role', ''),
                'EP2 Lat':              ep2.get('lat', ''),
                'EP2 Lng':              ep2.get('lng', ''),
                'EP2 Headwall':         ep2.get('headwall', ''),
                'EP2 Flange/End Sect.': ep2.get('flange_end_section', ''),
            })
            # Mid from LineString geometry
            gl = geom_line(pin)
            r['Mid Lat'] = gl.get('Mid Lat', '')
            r['Mid Lng'] = gl.get('Mid Lng', '')
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_sign(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'Road', 'FMSS Asset', 'Key Code', 'Speed Limit', 'Notes',
        'Latitude', 'Longitude',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'sign':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            r.update({
                'Road':        at.get('road', ''),
                'FMSS Asset':  at.get('fmss_asset', ''),
                'Key Code':    at.get('key_code', ''),
                'Speed Limit': at.get('speed_limit', ''),
                'Notes':       at.get('notes', ''),
                **geom_point(pin),
            })
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_mile_marker(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'Mile Label', 'Road', 'Type',
        'Latitude', 'Longitude',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'mile_marker':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            r.update({
                'Mile Label': at.get('mile_label', ''),
                'Road':       at.get('road', ''),
                'Type':       at.get('type', ''),
                **geom_point(pin),
            })
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_gate(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'Name', 'Short Name', 'Type', 'Road', 'Loc Name', 'Notes',
        'Latitude', 'Longitude',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'gate':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            r.update({
                'Name':       at.get('name', ''),
                'Short Name': at.get('short_name', ''),
                'Type':       at.get('type', ''),
                'Road':       at.get('road', ''),
                'Loc Name':   at.get('loc_name', ''),
                'Notes':      at.get('notes', ''),
                **geom_point(pin),
            })
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_bridge(sections):
    """Combines bridge (point) and bridge_line (linestring) onto one sheet."""
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID',
        'Kind', 'Source', 'Status', 'Reported In',
        'Stationing', 'Stationing (ft)',
        'Name', 'Short Name', 'Road', 'FMSS Loc', 'FMSS Asset', 'Notes',
        'Latitude', 'Longitude',          # point features
        'Start Lat', 'Start Lng',          # line features
        'Mid Lat', 'Mid Lng',
        'End Lat', 'End Lng',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] not in ('bridge', 'bridge_line'):
                continue
            r = {**base(sec, pin), 'Kind': pin['kind']}
            at = pin.get('attrs', {})
            r.update({
                'Name':       at.get('name', ''),
                'Short Name': at.get('short_name', ''),
                'Road':       at.get('road', ''),
                'FMSS Loc':   at.get('fmss_loc', ''),
                'FMSS Asset': at.get('fmss_asset', ''),
                'Notes':      at.get('notes', ''),
            })
            if pin['geometry'].get('type') == 'Point':
                r.update(geom_point(pin))
            else:
                r.update(geom_line(pin))
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


def rows_parking(sections):
    cols = [
        'Section', 'Section Name', 'Project Code', 'Pin ID',
        'Source', 'Status', 'Reported In', 'Stationing', 'Stationing (ft)',
        'Loc Name', 'Road', 'FMSS Loc', 'Type', 'Notes',
        'Latitude', 'Longitude',
    ]
    rows = []
    for sec in sections:
        for pin in sec['pins']:
            if pin['kind'] != 'parking':
                continue
            r = {**base(sec, pin)}
            at = pin.get('attrs', {})
            r.update({
                'Loc Name': at.get('loc_name', ''),
                'Road':     at.get('road', ''),
                'FMSS Loc': at.get('fmss_loc', ''),
                'Type':     at.get('type', ''),
                'Notes':    at.get('notes', ''),
                **geom_point(pin),
            })
            rows.append([r.get(c, '') for c in cols])
    return cols, rows


# ---------------------------------------------------------------------------
# Sheet config: (tab_title, tab_color_hex, row_builder_fn)
# ---------------------------------------------------------------------------
SHEET_DEFS = [
    ('Guardrails',      '2E75B6', rows_guardrail),
    ('Retaining Walls', '843C0C', rows_wall),
    ('Culverts',        '7030A0', rows_culvert),
    ('Signs',           'ED7D31', rows_sign),
    ('Mile Markers',    '70AD47', rows_mile_marker),
    ('Gates',           'C00000', rows_gate),
    ('Bridges',         '4472C4', rows_bridge),
    ('Parking',         'FFC000', rows_parking),
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def make_header_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def make_alt_fill(hex_color):
    """Very light tint (10 % opacity simulation) for alternating rows."""
    # Blend with white: take hex and lighten significantly
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    lr = min(255, r + int((255 - r) * 0.88))
    lg = min(255, g + int((255 - g) * 0.88))
    lb = min(255, b + int((255 - b) * 0.88))
    return PatternFill('solid', fgColor=f'{lr:02X}{lg:02X}{lb:02X}')

HEADER_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
CELL_FONT    = Font(name='Calibri', size=10)
WRAP_ALIGN   = Alignment(vertical='top', wrap_text=False)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN = Side(border_style='thin', color='D9D9D9')
CELL_BORDER = Border(bottom=THIN)


def write_sheet(wb, title, hex_color, cols, data_rows):
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = hex_color

    # Freeze header row
    ws.freeze_panes = 'A2'

    header_fill = make_header_fill(hex_color)
    alt_fill    = make_alt_fill(hex_color)

    # Write header
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = HEADER_FONT
        cell.fill      = header_fill
        cell.alignment = HEADER_ALIGN

    # Write data rows
    for ri, row in enumerate(data_rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = CELL_FONT
            cell.alignment = WRAP_ALIGN
            if fill:
                cell.fill = fill

    # Auto-width (approx — openpyxl has no built-in measure)
    col_widths = [len(str(c)) for c in cols]
    for row in data_rows:
        for ci, val in enumerate(row):
            col_widths[ci] = min(50, max(col_widths[ci], len(str(val)) if val not in (None, '') else 0))
    for ci, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(8, width + 2)

    # Row height for header
    ws.row_dimensions[1].height = 30

    return ws


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def write_summary(wb, sections, sheet_defs, counts):
    ws = wb.create_sheet(title='Summary', index=0)
    ws.sheet_properties.tabColor = '404040'
    ws.freeze_panes = 'A2'

    hdr_fill = make_header_fill('404040')
    cols = ['Feature Type', 'Sheet', 'Count', 'Sections Present']
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = HEADER_FONT; cell.fill = hdr_fill; cell.alignment = HEADER_ALIGN

    # Section breakdown per kind
    from collections import defaultdict
    sec_per_kind = defaultdict(set)
    for sec in sections:
        for pin in sec['pins']:
            sec_per_kind[pin['kind']].add(sec['id'])

    kind_map = {
        'Guardrails':      ['guardrail'],
        'Retaining Walls': ['wall'],
        'Culverts':        ['culvert'],
        'Signs':           ['sign'],
        'Mile Markers':    ['mile_marker'],
        'Gates':           ['gate'],
        'Bridges':         ['bridge', 'bridge_line'],
        'Parking':         ['parking'],
    }
    for ri, (title, _, _) in enumerate(sheet_defs, 2):
        kinds = kind_map[title]
        sec_ids = sorted(set().union(*(sec_per_kind[k] for k in kinds)))
        ws.cell(row=ri, column=1, value=title).font = CELL_FONT
        ws.cell(row=ri, column=2, value=title).font = CELL_FONT
        ws.cell(row=ri, column=3, value=counts.get(title, 0)).font = CELL_FONT
        ws.cell(row=ri, column=4, value=', '.join(sec_ids)).font = CELL_FONT

    # Totals row
    total_r = len(sheet_defs) + 2
    ws.cell(row=total_r, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=total_r, column=3, value=sum(counts.values())).font = Font(name='Calibri', bold=True, size=10)

    for ci in range(1, 5):
        ws.column_dimensions[get_column_letter(ci)].width = [22, 22, 10, 25][ci-1]
    ws.row_dimensions[1].height = 22


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    bundle_path = os.path.join(DATA, 'prewalk-bundle.json')
    print(f'Loading {bundle_path} ...')
    with open(bundle_path, encoding='utf-8') as f:
        bundle = json.load(f)
    sections = bundle['sections']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    counts = {}
    all_sheets = []
    for title, color, builder in SHEET_DEFS:
        cols, data_rows = builder(sections)
        counts[title] = len(data_rows)
        all_sheets.append((title, color, cols, data_rows))

    write_summary(wb, sections, SHEET_DEFS, counts)
    for title, color, cols, data_rows in all_sheets:
        write_sheet(wb, title, color, cols, data_rows)

    wb.save(OUT_XLS)
    print(f'\nWrote: {OUT_XLS}')
    print()
    print('Per-sheet counts:')
    for title, n in counts.items():
        print(f'  {title:<20} {n:>4} rows')
    print(f'  {"TOTAL":<20} {sum(counts.values()):>4} rows')


if __name__ == '__main__':
    main()
